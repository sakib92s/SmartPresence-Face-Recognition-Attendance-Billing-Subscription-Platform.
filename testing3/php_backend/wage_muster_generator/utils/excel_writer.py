"""
Excel Writer Module for Generating Wage Muster Excel File
With formulas and Indian payroll compliance
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from datetime import datetime

def generate_wage_muster_excel(wage_data, org_name, tender_number, month, 
                              contractor_name, work_location, output_path):
    """
    Generate formatted wage muster Excel file with formulas
    
    Args:
        wage_data (list): List of wage records
        org_name (str): Name of organization
        tender_number (str): Tender number
        month (str): Month (e.g., DEC-2025)
        contractor_name (str): Contractor/Company name
        work_location (str): Factory/Work location
        output_path (str): Path to save output Excel file
    """
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Wage Muster"
    
    # Define styles
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(horizontal='center', vertical='center')
    amount_alignment = Alignment(horizontal='right', vertical='center')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write header information
    ws.merge_cells('A1:N1')
    ws['A1'] = f"WAGE MUSTER FORM-B"
    ws['A1'].font = Font(name='Arial', size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:N2')
    ws['A2'] = f"{org_name.upper()}"
    ws['A2'].font = Font(name='Arial', size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A3:N3')
    ws['A3'] = f"Tender No: {tender_number} | Month: {month}"
    ws['A3'].font = Font(name='Arial', size=11)
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A4:N4')
    ws['A4'] = f"Contractor: {contractor_name} | Location: {work_location}"
    ws['A4'].font = Font(name='Arial', size=11)
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A5:N5')
    ws['A5'] = f"Generated on: {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}"
    ws['A5'].font = Font(name='Arial', size=10, italic=True)
    ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Leave one empty row
    current_row = 7
    
    # Define column headers (matching output.xlsx format)
    headers = [
        'SR.NO',
        'NAME OF EMPLOYEE',
        'DESIGNATION',
        'Working Days',
        'ATTN',
        'DW',
        'EARN WAGE',
        'EDLI',
        'BONUS @8.33',
        'EPF@ 12%',
        'ESIC- 0.75%',
        'PT',
        'NET PAY',
        'SIGN'
    ]
    
    # Write column headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    current_row += 1
    
    # Write wage data with formulas
    for idx, record in enumerate(wage_data, start=current_row):
        row_num = idx
        
        # Fixed values
        ws.cell(row=row_num, column=1, value=record['SR.NO'])  # A: SR.NO
        ws.cell(row=row_num, column=2, value=record['NAME OF EMPLOYEE'])  # B: NAME
        ws.cell(row=row_num, column=3, value=record['DESIGNATION'])  # C: DESIGNATION
        ws.cell(row=row_num, column=4, value=record['WORKING_DAYS'])  # D: Working Days
        ws.cell(row=row_num, column=5, value=record['ATTN'])  # E: ATTN
        
        # Formula columns
        # F: DW (already calculated, but show value)
        ws.cell(row=row_num, column=6, value=record['DW'])
        
        # G: EARN WAGE = E*F
        earn_wage_formula = f"=E{row_num}*F{row_num}"
        ws.cell(row=row_num, column=7, value=earn_wage_formula)
        
        # H: EDLI (value, but could be formula: =IF(G{row}>=15000,15000,G{row}))
        ws.cell(row=row_num, column=8, value=record['EDLI'])
        
        # I: BONUS = G*8.33% (ONLY FOR UNSKILLED)
        # We'll use conditional formula in Excel
        bonus_formula = f'=IF(C{row_num}="UNSKILLED", G{row_num}*0.0833, 0)'
        ws.cell(row=row_num, column=9, value=bonus_formula)
        
        # J: EPF = H*12%
        epf_formula = f"=H{row_num}*0.12"
        ws.cell(row=row_num, column=10, value=epf_formula)
        
        # K: ESIC = G*0.75% (ONLY FOR UNSKILLED)
        # Conditional formula in Excel
        esic_formula = f'=IF(C{row_num}="UNSKILLED", G{row_num}*0.0075, 0)'
        ws.cell(row=row_num, column=11, value=esic_formula)
        
        # L: PT (value based on gender rules - we'll put actual value since it's complex)
        # For Excel, we'll put the actual calculated value
        ws.cell(row=row_num, column=12, value=record['PT'])
        
        # M: NET PAY = (G+I)-(J+K+L)
        net_pay_formula = f"=(G{row_num}+I{row_num})-(J{row_num}+K{row_num}+L{row_num})"
        ws.cell(row=row_num, column=13, value=net_pay_formula)
        
        # N: SIGN (empty)
        ws.cell(row=row_num, column=14, value="")
        
        # Apply styles
        for col in range(1, 15):
            cell = ws.cell(row=row_num, column=col)
            cell.font = data_font
            cell.border = border
            
            # Apply right alignment for monetary columns
            if col >= 6:  # Columns F onwards are monetary
                cell.alignment = amount_alignment
                if col in [7, 8, 9, 10, 11, 12, 13]:  # Monetary value columns
                    cell.number_format = '#,##0.00'
            else:
                cell.alignment = data_alignment
    
    # Update current_row after writing data
    current_row = len(wage_data) + 8
    
    # Write totals row
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws.cell(row=current_row, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='right', vertical='center')
    
    # Total formulas
    total_formulas = {
        7: f"=SUM(G8:G{current_row-1})",  # G: Total EARN WAGE
        8: f"=SUM(H8:H{current_row-1})",  # H: Total EDLI
        9: f"=SUM(I8:I{current_row-1})",  # I: Total BONUS
        10: f"=SUM(J8:J{current_row-1})",  # J: Total EPF
        11: f"=SUM(K8:K{current_row-1})",  # K: Total ESIC
        12: f"=SUM(L8:L{current_row-1})",  # L: Total PT
        13: f"=SUM(M8:M{current_row-1})",  # M: Total NET PAY
    }
    
    for col, formula in total_formulas.items():
        cell = ws.cell(row=current_row, column=col, value=formula)
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = amount_alignment
        cell.number_format = '#,##0.00'
    
    # Write calculation rules and notes
    current_row += 2
    
    # Rules Section
    ws.merge_cells(f'A{current_row}:N{current_row}')
    ws.cell(row=current_row, column=1, value='CALCULATION RULES AND FORMULAS:').font = Font(bold=True, size=12)
    current_row += 1
    
    rules = [
        ('DAILY WAGE RATES:', 'Unskilled: ₹805, Semi-Skilled: ₹893, Skilled: ₹981, High Skilled: ₹1065'),
        ('EARNED WAGE (G):', '= ATTENDANCE (E) × DAILY WAGE (F)'),
        ('EDLI (H):', '= IF(EARNED WAGE >= 15000, 15000, EARNED WAGE)'),
        ('BONUS (I):', '= IF(DESIGNATION="UNSKILLED", EARNED WAGE (G) × 8.33%, 0)'),
        ('EPF (J):', '= EDLI (H) × 12% (Applicable to all categories)'),
        ('ESIC (K):', '= IF(DESIGNATION="UNSKILLED", EARNED WAGE (G) × 0.75%, 0)'),
        ('PROFESSIONAL TAX (L) - MALE:', 'If Earned Wage < ₹7500: ₹0; ₹7500-₹10000: ₹175; >₹10000: ₹200'),
        ('PROFESSIONAL TAX (L) - FEMALE:', 'If Earned Wage ≤ ₹25000: ₹0; >₹25000: ₹200'),
        ('NET PAY (M):', '= (EARNED WAGE + BONUS) - (EPF + ESIC + PT)'),
        ('VERIFICATION:', 'Column M should match: =(G+I)-(J+K+L)'),
        ('SPECIAL RULES:', 'Bonus and ESIC applicable ONLY for Unskilled category. Other categories get 0.')
    ]
    
    for rule_title, rule_desc in rules:
        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws.cell(row=current_row, column=1, value=rule_title).font = Font(bold=True)
        
        ws.merge_cells(f'D{current_row}:N{current_row}')
        ws.cell(row=current_row, column=4, value=rule_desc)
        
        current_row += 1
    
    # Statistics Section
    current_row += 1
    ws.merge_cells(f'A{current_row}:N{current_row}')
    ws.cell(row=current_row, column=1, value='SUMMARY STATISTICS:').font = Font(bold=True, size=12)
    current_row += 1
    
    stats = [
        ('Total Employees:', len(wage_data)),
        ('Unskilled Employees:', f'=COUNTIF(C8:C{current_row-8}, "UNSKILLED")'),
        ('Total Earned Wage:', f"=SUM(G8:G{current_row-8})"),
        ('Total Bonus (Unskilled Only):', f"=SUM(I8:I{current_row-8})"),
        ('Total Deductions:', f"=SUM(J8:L{current_row-8})"),
        ('Total Net Payable:', f"=SUM(M8:M{current_row-8})"),
        ('Average Net Pay per Employee:', f"=AVERAGE(M8:M{current_row-8})")
    ]
    
    for stat_title, stat_value in stats:
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws.cell(row=current_row, column=1, value=stat_title).font = Font(bold=True)
        
        ws.merge_cells(f'F{current_row}:N{current_row}')
        if isinstance(stat_value, str) and stat_value.startswith('='):
            ws.cell(row=current_row, column=6, value=stat_value)
        else:
            ws.cell(row=current_row, column=6, value=stat_value)
        ws.cell(row=current_row, column=6).alignment = amount_alignment
        if isinstance(stat_value, (int, float)) or (isinstance(stat_value, str) and stat_value.startswith('=')):
            ws.cell(row=current_row, column=6).number_format = '#,##0.00'
        
        current_row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                cell_value = str(cell.value)
                if cell_value.startswith('='):
                    # For formulas, consider the formula length
                    max_length = max(max_length, len(cell_value))
                else:
                    max_length = max(max_length, len(cell_value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze panes for headers
    ws.freeze_panes = 'A8'
    
    # Save workbook
    wb.save(output_path)

def create_wage_summary_sheet(wage_data, output_path):
    """
    Create a summary sheet with wage statistics
    
    Args:
        wage_data (list): List of wage records
        output_path (str): Path to save output Excel file
    """
    # Convert wage data to DataFrame
    df = pd.DataFrame(wage_data)
    
    # Create summary statistics
    summary = {
        'Total Employees': len(wage_data),
        'Unskilled Employees': len(df[df['DESIGNATION'] == 'UNSKILLED']),
        'Total Earned Wage': df['EARN_WAGE'].sum(),
        'Total Bonus': df['BONUS_8.33%'].sum(),
        'Total EPF': df['EPF_12%'].sum(),
        'Total ESIC': df['ESIC_0.75%'].sum(),
        'Total Professional Tax': df['PT'].sum(),
        'Total Net Pay': df['NET_PAY'].sum(),
        'Average Daily Wage': df['DW'].mean(),
        'Average Net Pay': df['NET_PAY'].mean()
    }
    
    return summary